#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
create_task_excel.py  –  Tao file Excel phan chia task do an TKM Cuoi Ky
pip install openpyxl  →  python3 create_task_excel.py
"""
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def fill(c): return PatternFill('solid', fgColor=c)
def border():
    s = Side(style='thin', color='AAAAAA')
    return Border(left=s, right=s, top=s, bottom=s)
def hdr(ws, r, c, v, bg='1F4E79', fg='FFFFFF', sz=10, bold=True):
    cell = ws.cell(r, c, v)
    cell.fill = fill(bg)
    cell.font = Font(bold=bold, color=fg, size=sz)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border()
def dat(ws, r, c, v, bg=None, bold=False, align='left'):
    cell = ws.cell(r, c, v)
    if bg: cell.fill = fill(bg)
    cell.font = Font(bold=bold, size=10)
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
    cell.border = border()

TASKS = [
    # (Phase, ID, Task, Description, File/Tool, Priority, Status, Note)
    ('Giai doan 1\nNghien cuu', 'T01', 'Doc de bai & xac dinh yeu cau',
     'Phan tich PDF de bai, liet ke cac yeu cau bat buoc', 'DeTaiCuoiKy_26_TKM.pdf', 'CRITICAL', 'DONE', ''),

    ('Giai doan 1\nNghien cuu', 'T02', 'Nghien cuu MPLS LDP co ban',
     'Label switching, LIB, FIB, LSP, Ingress/Transit/Egress LSR', 'RFC 3031, 5036', 'HIGH', 'DONE', ''),

    ('Giai doan 1\nNghien cuu', 'T03', 'Nghien cuu kien truc Metro Ethernet MAN',
     'E-Line, E-LAN, E-Tree, MPLS backbone, PE/P/CE roles', 'MEF standards', 'HIGH', 'DONE', ''),

    ('Giai doan 1\nNghien cuu', 'T04', 'Nghien cuu FRRouting (FRR)',
     'Cai dat, cau hinh OSPF, LDP, BGP tren Linux namespace', 'FRR docs', 'HIGH', 'DONE', ''),

    ('Giai doan 2\nThiet ke', 'T05', 'Thiet ke topology tong the',
     '3 chi nhanh + MPLS backbone PE1-P1-P2-PE2, ve so do', 'draw.io', 'CRITICAL', 'DONE', 'Xem DOC/Metro_MPLS_Topology.drawio'),

    ('Giai doan 2\nThiet ke', 'T06', 'Thiet ke dia chi IP',
     'Backbone 10.0.x.x/30, Loopback 10.255.0.x/32, CE-PE 172.16.x.x/30, LAN 192.168.x.x/24',
     'IP Plan', 'CRITICAL', 'DONE', 'Xem main_topology.py configure_ip()'),

    ('Giai doan 2\nThiet ke', 'T07', 'Thiet ke Site A – Flat Network',
     '1 switch (sA) + 3 host hA1/hA2/hA3, gateway CE1 192.168.10.1', 'main_topology.py', 'HIGH', 'DONE', ''),

    ('Giai doan 2\nThiet ke', 'T08', 'Thiet ke Site B – 3-Tier',
     'Core sw sBcore → Dist sBdist1/2 → Access sBacc1/2 → 4 hosts hB1-hB4', 'main_topology.py', 'HIGH', 'DONE', ''),

    ('Giai doan 2\nThiet ke', 'T09', 'Thiet ke Site C – Leaf-Spine',
     '2 Spine (spine1/2) + 3 Leaf (leaf1/2/3) full-mesh + 3 hosts hC1-hC3', 'main_topology.py', 'HIGH', 'DONE', ''),

    ('Giai doan 3\nTrien khai', 'T10', 'Lap trinh topology Mininet',
     'Class LinuxRouter, build_topology(), TCLink, OVSKernelSwitch', 'main_topology.py', 'CRITICAL', 'DONE', 'sudo python3 main_topology.py'),

    ('Giai doan 3\nTrien khai', 'T11', 'Cau hinh IP tat ca interface',
     'configure_ip() – loopback, backbone, WAN, LAN, MPLS sysctl', 'main_topology.py', 'CRITICAL', 'DONE', ''),

    ('Giai doan 3\nTrien khai', 'T12', 'Cau hinh FRR – P routers (OSPF+LDP)',
     'p1 & p2: OSPF area 0, LDP, mpls enable tren tat ca eth', 'main_topology.py FRR_CONFIGS', 'CRITICAL', 'DONE', ''),

    ('Giai doan 3\nTrien khai', 'T13', 'Cau hinh FRR – PE routers (OSPF+LDP+static)',
     'pe1 & pe2: OSPF redistribute, LDP, static route den CE subnets', 'main_topology.py FRR_CONFIGS', 'CRITICAL', 'DONE', ''),

    ('Giai doan 3\nTrien khai', 'T14', 'Cau hinh FRR – CE routers (static routes)',
     'ce1/ce2/ce3: default route qua PE, cross-site static routes', 'main_topology.py FRR_CONFIGS', 'CRITICAL', 'DONE', ''),

    ('Giai doan 3\nTrien khai', 'T15', 'Khoi dong FRR daemons tu dong',
     'start_frr(): zebra → ospfd → ldpd, cho convergence 20s', 'main_topology.py', 'CRITICAL', 'DONE', ''),

    ('Giai doan 4\nKiem thu', 'T16', 'Kiem tra ket noi noi bo tung site',
     'hA1 ping hA2, hB1 ping hB2, hC1 ping hC2', 'Mininet CLI', 'HIGH', 'TODO', 'Chay sau khi khoi dong mang'),

    ('Giai doan 4\nKiem thu', 'T17', 'Kiem tra ket noi cross-site qua MPLS',
     'hA1 ping hB1, hA1 ping hC1, hB1 ping hC1', 'Mininet CLI', 'CRITICAL', 'TODO', ''),

    ('Giai doan 4\nKiem thu', 'T18', 'Xac nhan MPLS LDP hoat dong',
     'pe1 vtysh "show mpls ldp binding", "show mpls ldp neighbor"', 'vtysh', 'CRITICAL', 'TODO', ''),

    ('Giai doan 4\nKiem thu', 'T19', 'Xac nhan OSPF neighbor',
     'pe1 vtysh "show ip ospf neighbor" → phai thay p1', 'vtysh', 'HIGH', 'TODO', ''),

    ('Giai doan 4\nKiem thu', 'T20', 'Xem MPLS forwarding table tren P router',
     'p1 vtysh "show mpls forwarding-table"', 'vtysh', 'HIGH', 'TODO', ''),

    ('Giai doan 5\nDo luong', 'T21', 'Do Ping – RTT & Packet Loss',
     'run_ping() giua 3 cap site: A↔B, A↔C, B↔C (30 goi)', 'performance_test.py', 'CRITICAL', 'TODO', ''),

    ('Giai doan 5\nDo luong', 'T22', 'Do iPerf TCP – Throughput',
     'run_iperf() TCP 10s giua 3 cap, ghi bandwidth Mbps', 'performance_test.py', 'CRITICAL', 'TODO', ''),

    ('Giai doan 5\nDo luong', 'T23', 'Do iPerf UDP – Jitter & Loss',
     'run_iperf() UDP 50Mbps 10s, ghi jitter ms va loss %', 'performance_test.py', 'HIGH', 'TODO', ''),

    ('Giai doan 5\nDo luong', 'T24', 'Traceroute – xac nhan duong di qua MPLS',
     'hA1 traceroute hC1 → phai di qua CE1→PE1→P1→P2→PE2→CE3', 'performance_test.py', 'HIGH', 'TODO', ''),

    ('Giai doan 6\nBao cao', 'T25', 'Tao bao cao Excel hieu nang',
     'generate_excel_report.py → 4 sheet: Ping, Throughput, MPLS, Comparison', 'generate_excel_report.py', 'HIGH', 'TODO', 'python3 generate_excel_report.py'),

    ('Giai doan 6\nBao cao', 'T26', 'Phan tich so sanh 3 kien truc LAN',
     'Flat vs 3-Tier vs Leaf-Spine: latency, throughput, scalability, fault tolerance', 'generate_excel_report.py Sheet4', 'HIGH', 'TODO', ''),

    ('Giai doan 6\nBao cao', 'T27', 'Viet bao cao Word / LaTeX',
     'Chuong 1: Ly thuyet, Chuong 2: Thiet ke, Chuong 3: Trien khai, Chuong 4: Ket qua', 'docs/', 'CRITICAL', 'TODO', ''),

    ('Giai doan 6\nBao cao', 'T28', 'Chup screenshot / ghi video demo',
     'Quay man hinh chay topology, ping, traceroute, vtysh commands', 'OBS / recordmydesktop', 'MEDIUM', 'TODO', ''),
]

PRIORITY_COLORS = {
    'CRITICAL': 'E74C3C',
    'HIGH':     'F39C12',
    'MEDIUM':   '27AE60',
    'LOW':      '2980B9',
}
STATUS_COLORS = {
    'DONE':        'D5F5E3',
    'IN_PROGRESS': 'FCF3CF',
    'TODO':        'FADBD8',
}
PHASE_COLORS = {
    'Giai doan 1\nNghien cuu': 'D6EAF8',
    'Giai doan 2\nThiet ke':   'D5F5E3',
    'Giai doan 3\nTrien khai': 'FDEBD0',
    'Giai doan 4\nKiem thu':   'E8DAEF',
    'Giai doan 5\nDo luong':   'FDFEFE',
    'Giai doan 6\nBao cao':    'FDEDEC',
}

def create_task_sheet(wb):
    ws = wb.create_sheet('Task List - Do An TKM')
    ws.sheet_view.showGridLines = False

    # Column widths
    widths = [20, 6, 32, 46, 28, 10, 14, 24]
    cols   = ['A','B','C','D','E','F','G','H']
    for c, w in zip(cols, widths):
        ws.column_dimensions[c].width = w

    # Title
    ws.merge_cells('A1:H1')
    t = ws['A1']
    t.value = '📋 KE HOACH THUC HIEN DO AN CUOI KY – TKM – MSSV: 52300267'
    t.font = Font(bold=True, size=15, color='FFFFFF')
    t.fill = fill('154360')
    t.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 38

    # Sub-title
    ws.merge_cells('A2:H2')
    s = ws['A2']
    s.value = 'De tai: Thiet ke va trien khai mang Metro Ethernet su dung MPLS cho ket noi da chi nhanh doanh nghiep'
    s.font = Font(italic=True, size=11, color='2C3E50')
    s.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 22

    # Header row
    headers = ['Giai Doan', 'ID', 'Nhiem Vu', 'Mo Ta Chi Tiet', 'File / Cong Cu', 'Uu Tien', 'Trang Thai', 'Ghi Chu']
    for i, h in enumerate(headers, 1):
        hdr(ws, 3, i, h, bg='1A5276', sz=11)
    ws.row_dimensions[3].height = 26

    # Data rows
    for idx, (phase, tid, task, desc, tool, prio, status, note) in enumerate(TASKS):
        row = idx + 4
        phase_bg = PHASE_COLORS.get(phase, 'FFFFFF')
        dat(ws, row, 1, phase, bg=phase_bg, bold=True, align='center')
        dat(ws, row, 2, tid, bg=phase_bg, bold=True, align='center')
        dat(ws, row, 3, task, bg=phase_bg, bold=False)
        dat(ws, row, 4, desc, bg=phase_bg)
        dat(ws, row, 5, tool, bg=phase_bg)

        # Priority cell
        pc = ws.cell(row, 6, prio)
        pc.fill = fill(PRIORITY_COLORS.get(prio, 'FFFFFF'))
        pc.font = Font(bold=True, size=9, color='FFFFFF')
        pc.alignment = Alignment(horizontal='center', vertical='center')
        pc.border = border()

        # Status cell
        sc = ws.cell(row, 7, status)
        sc.fill = fill(STATUS_COLORS.get(status, 'FFFFFF'))
        sc.font = Font(bold=True, size=9)
        sc.alignment = Alignment(horizontal='center', vertical='center')
        sc.border = border()

        dat(ws, row, 8, note, bg=phase_bg)
        ws.row_dimensions[row].height = 30

    # Legend
    legend_row = len(TASKS) + 6
    ws.merge_cells(f'A{legend_row}:H{legend_row}')
    lg = ws[f'A{legend_row}']
    lg.value = 'CHU THICH: CRITICAL=Bat buoc  |  HIGH=Quan trong  |  MEDIUM=Khuyen nghi  |  DONE=Hoan thanh  |  TODO=Chua lam'
    lg.font = Font(italic=True, size=10, color='555555')
    lg.alignment = Alignment(horizontal='center')

    return ws


def create_summary_sheet(wb):
    ws = wb.create_sheet('Tom Tat')
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 50

    ws.merge_cells('A1:B1')
    t = ws['A1']
    t.value = 'TOM TAT DO AN CUOI KY - TKM'
    t.font = Font(bold=True, size=14, color='FFFFFF')
    t.fill = fill('1A252F')
    t.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 32

    items = [
        ('Ten de tai', 'Thiet ke va trien khai mang Metro Ethernet su dung MPLS'),
        ('MSSV', '52300267'),
        ('Truong', 'Dai hoc Ton Duc Thang'),
        ('Nganh', 'Mang may tinh va truyen thong du lieu'),
        ('Cong nghe chinh', 'Mininet + FRRouting (OSPF + LDP MPLS)'),
        ('Ngon ngu', 'Python 3'),
        ('Kien truc backbone', 'PE1 -- P1 -- P2 -- PE2 (MPLS LDP)'),
        ('Site A', 'Flat Network – 1 switch – 192.168.10.0/24'),
        ('Site B', '3-Tier (Core-Dist-Access) – 5 switch – 192.168.20.0/24'),
        ('Site C', 'Leaf-Spine – 5 switch – 192.168.30.0/24'),
        ('File chinh', 'main_topology.py → sudo python3 main_topology.py'),
        ('Do hieu nang', 'performance_test.py → full_test(net) trong CLI'),
        ('Bao cao Excel', 'generate_excel_report.py → python3 generate_excel_report.py'),
        ('Tao task list', 'create_task_excel.py (file nay)'),
        ('Tong so task', str(len(TASKS))),
        ('Task DONE', str(sum(1 for t in TASKS if t[6]=='DONE'))),
        ('Task TODO', str(sum(1 for t in TASKS if t[6]=='TODO'))),
    ]

    for idx, (k, v) in enumerate(items):
        r = idx + 2
        dat(ws, r, 1, k, bg='D6EAF8' if idx%2==0 else 'EBF5FB', bold=True)
        dat(ws, r, 2, v, bg='D6EAF8' if idx%2==0 else 'EBF5FB')
        ws.row_dimensions[r].height = 20
    return ws


def main():
    wb = Workbook()
    wb.remove(wb.active)
    create_summary_sheet(wb)
    create_task_sheet(wb)

    os.makedirs('results', exist_ok=True)
    out = 'results/task_list_tkm_52300267.xlsx'
    wb.save(out)
    print(f'[OK] Task Excel da tao: {out}')
    print(f'     Tong: {len(TASKS)} task | DONE: {sum(1 for t in TASKS if t[6]=="DONE")} | TODO: {sum(1 for t in TASKS if t[6]=="TODO")}')

if __name__ == '__main__':
    main()
