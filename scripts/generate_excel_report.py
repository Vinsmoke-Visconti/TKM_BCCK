#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
generate_excel_report.py
------------------------
Tao file Excel bao cao hieu nang tu ket qua JSON cua performance_test.py.

Cai dat: pip install openpyxl
Chay   : python3 generate_excel_report.py [report_json_path]
"""

import json
import sys
import os
import glob
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.chart.series import DataLabelList
except ImportError:
    print('[ERROR] Can cai openpyxl: pip install openpyxl')
    sys.exit(1)


# ---------- STYLE HELPERS ----------
def header_fill(color='1F4E79'):
    return PatternFill('solid', fgColor=color)

def cell_fill(color):
    return PatternFill('solid', fgColor=color)

def thin_border():
    s = Side(style='thin', color='999999')
    return Border(left=s, right=s, top=s, bottom=s)

def set_header(ws, row, col, value, color='1F4E79', font_color='FFFFFF', size=11, bold=True):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = header_fill(color)
    c.font = Font(bold=bold, color=font_color, size=size)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = thin_border()
    return c

def set_data(ws, row, col, value, fill_color=None, bold=False, align='center', num_format=None):
    c = ws.cell(row=row, column=col, value=value)
    if fill_color:
        c.fill = cell_fill(fill_color)
    c.font = Font(bold=bold, size=10)
    c.alignment = Alignment(horizontal=align, vertical='center')
    c.border = thin_border()
    if num_format:
        c.number_format = num_format
    return c


# ---------- SHEET 1: Ping Results ----------
def sheet_ping(wb, data):
    ws = wb.create_sheet('Ping - Delay & Loss')
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 16
    for col in ['B','C','D','E','F']:
        ws.column_dimensions[col].width = 14

    # Title
    ws.merge_cells('A1:F1')
    c = ws['A1']
    c.value = 'KET QUA PING - DO DO TRE & TI LE MAT GOI'
    c.font = Font(bold=True, size=14, color='FFFFFF')
    c.fill = header_fill('154360')
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # Header
    headers = ['Huong', 'RTT Min (ms)', 'RTT Avg (ms)', 'RTT Max (ms)', 'Jitter (ms)', 'Packet Loss (%)']
    for i, h in enumerate(headers, 1):
        set_header(ws, 2, i, h, color='1A5276')
    ws.row_dimensions[2].height = 22

    colors = ['EBF5FB', 'D6EAF8', 'EBF5FB']
    for idx, r in enumerate(data):
        row = idx + 3
        fill = colors[idx % len(colors)]
        set_data(ws, row, 1, r.get('label', r['src']), fill_color=fill, align='left')
        set_data(ws, row, 2, r['rtt_min'], fill_color=fill)
        set_data(ws, row, 3, r['rtt_avg'], fill_color=fill, bold=True)
        set_data(ws, row, 4, r['rtt_max'], fill_color=fill)
        set_data(ws, row, 5, r['rtt_mdev'], fill_color=fill)
        loss_color = 'FADBD8' if r['loss_pct'] > 0 else 'D5F5E3'
        set_data(ws, row, 6, r['loss_pct'], fill_color=loss_color, bold=True)

    # Chart - RTT Avg
    chart_row = len(data) + 4
    chart = BarChart()
    chart.type = 'col'
    chart.title = 'RTT Trung Binh giua cac chi nhanh (ms)'
    chart.y_axis.title = 'RTT (ms)'
    chart.x_axis.title = 'Huong'
    chart.style = 10

    data_ref = Reference(ws, min_col=4, max_col=4,
                         min_row=2, max_row=2 + len(data))
    cats = Reference(ws, min_col=1, min_row=3, max_row=2 + len(data))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    chart.width = 18
    chart.height = 11
    ws.add_chart(chart, f'A{chart_row}')
    return ws


# ---------- SHEET 2: Throughput ----------
def sheet_throughput(wb, tcp_data, udp_data):
    ws = wb.create_sheet('iPerf - Throughput')
    ws.sheet_view.showGridLines = False

    ws.column_dimensions['A'].width = 16
    for col in ['B','C','D','E']:
        ws.column_dimensions[col].width = 16

    # TCP Section
    ws.merge_cells('A1:C1')
    c = ws['A1']
    c.value = 'IPERF TCP - THROUGHPUT'
    c.font = Font(bold=True, size=13, color='FFFFFF')
    c.fill = header_fill('1E8449')
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 26

    for i, h in enumerate(['Huong', 'Che do', 'Throughput (Mbps)'], 1):
        set_header(ws, 2, i, h, color='196F3D')
    ws.row_dimensions[2].height = 22

    for idx, r in enumerate(tcp_data):
        row = idx + 3
        fill = 'EAFAF1' if idx % 2 == 0 else 'D5F5E3'
        set_data(ws, row, 1, r.get('label', r['src']), fill_color=fill, align='left')
        set_data(ws, row, 2, 'TCP', fill_color=fill)
        bw = r['bandwidth_mbps'] if r['bandwidth_mbps'] else 0
        set_data(ws, row, 3, bw, fill_color=fill, bold=True)

    # UDP Section
    udp_start = len(tcp_data) + 5
    ws.merge_cells(f'A{udp_start}:E{udp_start}')
    c = ws[f'A{udp_start}']
    c.value = 'IPERF UDP - JITTER & PACKET LOSS'
    c.font = Font(bold=True, size=13, color='FFFFFF')
    c.fill = header_fill('7D6608')
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[udp_start].height = 26

    for i, h in enumerate(['Huong', 'Che do', 'Bandwidth (Mbps)', 'Jitter (ms)', 'Loss (%)'], 1):
        set_header(ws, udp_start + 1, i, h, color='9A7D0A')
    ws.row_dimensions[udp_start + 1].height = 22

    for idx, r in enumerate(udp_data):
        row = udp_start + 2 + idx
        fill = 'FEFBD8' if idx % 2 == 0 else 'FCF3CF'
        set_data(ws, row, 1, r.get('label', r['src']), fill_color=fill, align='left')
        set_data(ws, row, 2, 'UDP', fill_color=fill)
        bw = r['bandwidth_mbps'] if r['bandwidth_mbps'] else 0
        set_data(ws, row, 3, bw, fill_color=fill, bold=True)
        jit = r['jitter_ms'] if r['jitter_ms'] else 0
        set_data(ws, row, 4, jit, fill_color=fill)
        ls_color = 'FADBD8' if (r['loss_pct'] or 0) > 1 else fill
        set_data(ws, row, 5, r['loss_pct'] or 0, fill_color=ls_color)

    # Chart
    chart_row = udp_start + len(udp_data) + 4
    chart = BarChart()
    chart.type = 'col'
    chart.title = 'Throughput TCP giua cac chi nhanh (Mbps)'
    chart.y_axis.title = 'Mbps'
    chart.style = 26
    data_ref = Reference(ws, min_col=3, max_col=3, min_row=2, max_row=2 + len(tcp_data))
    cats = Reference(ws, min_col=1, min_row=3, max_row=2 + len(tcp_data))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 18
    chart.height = 11
    ws.add_chart(chart, f'A{chart_row}')
    return ws


# ---------- SHEET 3: MPLS Analysis ----------
def sheet_mpls(wb):
    ws = wb.create_sheet('MPLS Analysis')
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 30

    ws.merge_cells('A1:C1')
    c = ws['A1']
    c.value = 'PHAN TICH HOAT DONG MPLS LABEL SWITCHING'
    c.font = Font(bold=True, size=14, color='FFFFFF')
    c.fill = header_fill('512E5F')
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    items = [
        ('Cong nghe', 'MPLS (Multi-Protocol Label Switching)', 'Chuyen mach nhan da giao thuc'),
        ('Protocol IGP', 'OSPFv2 - Area 0 (Backbone)', 'Dinh tuyen noi bo trong ISP'),
        ('Protocol Label', 'LDP (Label Distribution Protocol)', 'Phan phoi nhan tu dong'),
        ('Router P1 Role', 'Provider Core - Transit LSR', 'Chuyen tiep goi dua tren nhan'),
        ('Router P2 Role', 'Provider Core - Transit LSR', 'Chuyen tiep goi dua tren nhan'),
        ('Router PE1 Role', 'Provider Edge - Ingress/Egress LSR', 'Push/Pop/Swap label cho Site A, B'),
        ('Router PE2 Role', 'Provider Edge - Ingress/Egress LSR', 'Push/Pop/Swap label cho Site C'),
        ('Site A Subnet', '192.168.10.0/24 - Flat Network', 'CE1 gateway: 192.168.10.1'),
        ('Site B Subnet', '192.168.20.0/24 - 3-Tier', 'CE2 gateway: 192.168.20.1'),
        ('Site C Subnet', '192.168.30.0/24 - Leaf-Spine', 'CE3 gateway: 192.168.30.1'),
        ('Backbone Subnet', '10.0.12.0/30, 10.0.13.0/30, 10.0.24.0/30', 'P-PE interconnects'),
        ('CE-PE WAN', '172.16.1.0/30, 172.16.2.0/30, 172.16.3.0/30', 'Customer uplinks'),
        ('Loopbacks', '10.255.0.1-4/32', 'Router-ID cho OSPF & LDP'),
        ('Label Range', '16 - 100000', 'MPLS label platform_labels'),
        ('QoS WAN', 'bw=100Mbps, delay=5ms (CE-PE)', 'Mo phong SLA thuc te'),
        ('QoS Core', 'bw=1000Mbps, delay=1ms (P-PE)', 'Cap quang loi ISP'),
        ('STP', 'Disabled tren OVS switches', 'Tranh blocking trong lab'),
    ]

    headers = ['Thong so', 'Gia tri', 'Mo ta']
    for i, h in enumerate(headers, 1):
        set_header(ws, 2, i, h, color='6C3483')
    ws.row_dimensions[2].height = 22

    for idx, (k, v, d) in enumerate(items):
        row = idx + 3
        fill = 'F5EEF8' if idx % 2 == 0 else 'EBD8F5'
        set_data(ws, row, 1, k, fill_color=fill, bold=True, align='left')
        set_data(ws, row, 2, v, fill_color=fill, align='left')
        set_data(ws, row, 3, d, fill_color=fill, align='left')
        ws.row_dimensions[row].height = 18
    return ws


# ---------- SHEET 4: Comparison ----------
def sheet_comparison(wb):
    ws = wb.create_sheet('So sanh kien truc LAN')
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20

    ws.merge_cells('A1:D1')
    c = ws['A1']
    c.value = 'SO SANH HIEU NANG - 3 KIEN TRUC MANG LAN'
    c.font = Font(bold=True, size=14, color='FFFFFF')
    c.fill = header_fill('1A252F')
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    headers = ['Tieu chi', 'Flat (Site A)', '3-Tier (Site B)', 'Leaf-Spine (Site C)']
    for i, h in enumerate(headers, 1):
        set_header(ws, 2, i, h, color='2C3E50')
    ws.row_dimensions[2].height = 22

    rows = [
        ('So lop switch', '1 lop', '3 lop (Core-Dist-Access)', '2 lop (Spine-Leaf)'),
        ('So switch', '1 SW', '5 SW', '5 SW (2 Spine + 3 Leaf)'),
        ('So hop L2', '1 hop', '3 hop', '2 hop'),
        ('Scalability', 'Thap - bi gioi han', 'Trung binh - VLAN', 'Cao - ECMP, Scale-Out'),
        ('Fault Tolerance', 'Thap - single point', 'Cao - redundant link', 'Rat cao - full mesh'),
        ('Do tre noi bo (du kien)', 'Thap nhat (~0.1ms)', 'Trung binh (~0.3ms)', 'Thap (~0.2ms)'),
        ('Throughput noi bo', 'Line-rate', 'Phu thuoc QoS', 'Line-rate, ECMP'),
        ('Phu hop voi', 'SME, van phong nho', 'Doanh nghiep trung binh', 'Data Center, doanh nghiep lon'),
        ('Chi phi trien khai', 'Thap', 'Trung binh', 'Cao'),
        ('Quan ly', 'Don gian', 'Phuc tap hon (VLAN)', 'Phuc tap (BGP/OSPF)'),
        ('Chuan cong nghe', 'IEEE 802.1D', 'IEEE 802.1Q VLAN', 'BGP ECMP / OSPF ECMP'),
        ('Anh huong MPLS', 'Nho (it hop)', 'Trung binh', 'Nho (parallel paths)'),
    ]

    fills = [
        ('D6EAF8', 'D5F5E3', 'FDEBD0'),
        ('EBF5FB', 'EAFAF1', 'FEF9E7'),
    ]
    for idx, (crit, a, b, c_val) in enumerate(rows):
        row = idx + 3
        fa, fb, fc = fills[idx % 2]
        set_data(ws, row, 1, crit, fill_color='F2F3F4', bold=True, align='left')
        set_data(ws, row, 2, a, fill_color=fa, align='center')
        set_data(ws, row, 3, b, fill_color=fb, align='center')
        set_data(ws, row, 4, c_val, fill_color=fc, align='center')
        ws.row_dimensions[row].height = 20
    return ws


# ---------- MAIN ----------
def create_report(report_path=None):
    # Tim file JSON moi nhat neu khong chi dinh
    if not report_path:
        files = glob.glob('/tmp/tkm_results/full_report_*.json')
        if not files:
            print('[WARN] Khong tim thay file ket qua. Tao bao cao mau...')
            report = {
                'timestamp': datetime.now().strftime('%Y%m%d_%H%M%S'),
                'ping_results': [
                    {'label': 'A->B', 'src': 'hA1', 'dst': '192.168.20.11',
                     'rtt_min': 10.2, 'rtt_avg': 12.5, 'rtt_max': 18.3, 'rtt_mdev': 1.8, 'loss_pct': 0.0},
                    {'label': 'A->C', 'src': 'hA1', 'dst': '192.168.30.11',
                     'rtt_min': 10.8, 'rtt_avg': 13.1, 'rtt_max': 19.7, 'rtt_mdev': 2.1, 'loss_pct': 0.0},
                    {'label': 'B->C', 'src': 'hB1', 'dst': '192.168.30.11',
                     'rtt_min': 10.5, 'rtt_avg': 12.9, 'rtt_max': 18.9, 'rtt_mdev': 1.9, 'loss_pct': 0.0},
                ],
                'iperf_tcp': [
                    {'label': 'B->A TCP', 'src': 'hB1', 'dst': 'hA1', 'bandwidth_mbps': 94.3},
                    {'label': 'C->A TCP', 'src': 'hC1', 'dst': 'hA1', 'bandwidth_mbps': 91.7},
                    {'label': 'C->B TCP', 'src': 'hC1', 'dst': 'hB1', 'bandwidth_mbps': 92.5},
                ],
                'iperf_udp': [
                    {'label': 'B->A UDP', 'src': 'hB1', 'dst': 'hA1',
                     'bandwidth_mbps': 49.8, 'jitter_ms': 0.23, 'loss_pct': 0.1},
                    {'label': 'C->A UDP', 'src': 'hC1', 'dst': 'hA1',
                     'bandwidth_mbps': 49.5, 'jitter_ms': 0.31, 'loss_pct': 0.2},
                ],
                'traceroutes': {},
            }
        else:
            report_path = sorted(files)[-1]
            with open(report_path) as f:
                report = json.load(f)
    else:
        with open(report_path) as f:
            report = json.load(f)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Xoa sheet mac dinh

    sheet_ping(wb, report['ping_results'])
    sheet_throughput(wb, report['iperf_tcp'], report['iperf_udp'])
    sheet_mpls(wb)
    sheet_comparison(wb)

    out_path = f'results/performance_report_{report["timestamp"]}.xlsx'
    os.makedirs('results', exist_ok=True)
    wb.save(out_path)
    print(f'\n[OK] Bao cao Excel da duoc tao: {out_path}')
    return out_path


if __name__ == '__main__':
    path = sys.argv[1] if len(sys.argv) > 1 else None
    create_report(path)
